from random import choice
from openpyxl import Workbook, load_workbook

'''
This script requires an excel file to be created with all of the student's names and genders
Check that the variable studentsDir contains the path of the student name excel file

    The first column should include each student's first and last name
    The second column should include each students gender, marked with M for male or F for female

    The first row of the spreadsheet can include a header or be blank.
    The script begins reading from the second row.

The script also requires a separate excel workbook for the past teams to check
Check that the pastTeamsDir contains the path of the past teams excel workbook

    The pastTeams.xlsx file is fairly simple and only requires one column
    The column should contain teams of students in the past.
    Place the first and last names of each student in a team seperated by ' & '
    Ex: Jordan Cottle & Sarah Woolley

    You can add headers and other blank rows for formatting, as long as they don't contain the ' & ' sequence

When the script is run, it will add the teams to the past teams workbook at the bottom of the first column.
It will also create a new excel workbook named 'teams.xlsx' (or overwrite it if 'teams.xlsx' exists) with the new teams

See the studentsExample.xlsx file for proper formatting of the students workbook
See the pastTeamsExample.xlsx file for proper formatting of the past teams workbook
See the teamsExample.xlsx file for the format of the expected output of the script

You can change the studentsDir, pastTeamsDir, and outputDir variables to point to the example files to test the script
'''

studentsDir = 'students.xlsx'
pastTeamsDir = 'pastTeams.xlsx'
outputDir = 'teams.xlsx'


class Student:
    def __init__(self, firstName, lastName):
        self.firstName = firstName
        self.lastName = lastName

    def __str__(self):
        return f'{self.firstName} {self.lastName}'

    def __eq__(self, value):
        if type(value) != type(self):
            raise NotImplementedError(f"Cannot compare {type(value)} to {type(self)}")

        return str(self) == str(value)


class Team:
    def __init__(self, students):
        self.members = []
        for student in students:
            self.members.append(student)

    def addMember(self, student):
        self.members.append(student)

    def __str__(self):
        return ' & '.join((str(member) for member in self.members))

    def __hash__(self):
        return hash(str(self))

    def __eq__(self, value):
        if type(value) != type(self):
            raise NotImplementedError(f"Cannot compare {type(value)} to {type(self)}")

        return str(self) == str(value)


def chooseTwo(items):
    a = choice(items)
    b = choice(items)
    while(b == a):
        b = choice(items)

    return a, b


def makeTeams(students, pastTeams):
    teams = []
    while len(students) > 1:
        team = Team(chooseTwo(students))

        while team in pastTeams:
            print(f'{team} have already been a team!')
            team = Team(chooseTwo(students))

        for student in team.members:
            students.remove(student)

        teams.append(team)
        pastTeams.add(team)

    return teams


# Get students from spreadsheet
studentWorkbook = load_workbook(studentsDir, read_only=True)
studentWorksheet = studentWorkbook.active

males = []
females = []
# skip header on row 1
for i in range(2, studentWorksheet.max_row + 1):
    firstName, lastName = studentWorksheet[f'A{i}'].value.split(" ", 1)
    gender = studentWorksheet[f'B{i}'].value

    student = Student(firstName, lastName)

    if gender.lower() == 'm':
        males.append(student)
    else:
        females.append(student)

# Get past teams from spreadsheet
pastTeamWorkbook = load_workbook(pastTeamsDir)
pastTeamWorksheet = pastTeamWorkbook.active

pastTeams = set()
for pastRow in range(1, pastTeamWorksheet.max_row+1):
    teamName = pastTeamWorksheet[f'A{pastRow}'].value

    if(teamName is None):
        print(f'Row {pastRow} is empty')
        continue

    memberNames = teamName.split(' & ')

    # Skip empty and single member teams as well as headers that don't split on the & symbol
    if(len(memberNames) < 2):
        continue

    members = []
    for memberName in memberNames:
        first, last = memberName.split(' ', 1)
        members.append(Student(first, last))

    team = Team(members)

    pastTeams.add(team)

# For new teams, avoid repeating a pair from a previous lab
maleTeams = makeTeams(males, pastTeams)
femaleTeams = makeTeams(females, pastTeams)

malesRemaining = len(males)
femalesRemaining = len(females)
remaining = malesRemaining + femalesRemaining

# handle extra students
if remaining == 1:  # odd number in class

    # Add extra male/female into existing male/female team (a single team of 3)
    if malesRemaining == 1:
        print('Creating male team of 3')
        choice(maleTeams).addMember(males[0])
    elif femalesRemaining == 1:
        print('Creating female team of 3')
        print(females[0])
        choice(femaleTeams).addMember(females[0])

elif remaining == 2:  # odd number of both males and females
    print('Creating mixed team')
    femaleTeams.append(Team([
        males[0],
        females[0]
    ]))

teams = maleTeams + femaleTeams

workbook = Workbook()

sheet = workbook.active

col = 'A'
row = 1
pastRow += 2  # Leave a blank line to separate groups of teams
for team in teams:

    # set names into cell
    sheet[f'{col}{row}'].value = str(team)
    pastTeamWorksheet[f'A{pastRow}'].value = str(team)

    row += 1
    pastRow += 1

workbook.save(outputDir)
pastTeamWorkbook.save(pastTeamsDir)
